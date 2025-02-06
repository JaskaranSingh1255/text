import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference, PieChart

# Load the processed data
file_path = "MobileServices1.xlsx"
df = pd.read_excel(file_path)

# Count occurrences of each category and subcategory
category_counts = df["Category"].value_counts()
subcategory_counts = df["Subcategory"].value_counts()
sub_subcategory_counts = df["Sub-Subcategory"].value_counts()

# Save aggregated data to a new sheet in the same file
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
    category_counts.to_frame("Count").to_excel(writer, sheet_name="Category Count")
    subcategory_counts.to_frame("Count").to_excel(writer, sheet_name="Subcategory Count")
    sub_subcategory_counts.to_frame("Count").to_excel(writer, sheet_name="Sub-Subcategory Count")

# Load workbook and select the active sheet
wb = openpyxl.load_workbook(file_path)

# Function to create a bar chart
def create_bar_chart(sheet, data_range, title, cell_position):
    chart = BarChart()
    chart.title = title
    chart.x_axis.title = "Categories"
    chart.y_axis.title = "Count"
    chart.add_data(data_range, titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=data_range.max_row))
    sheet.add_chart(chart, cell_position)

# Function to create a pie chart
def create_pie_chart(sheet, data_range, title, cell_position):
    chart = PieChart()
    chart.title = title
    chart.add_data(data_range, titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=data_range.max_row))
    sheet.add_chart(chart, cell_position)

# Add bar charts for subcategories and sub-subcategories
sheet_subcategory = wb["Subcategory Count"]
data_range_subcategory = Reference(sheet_subcategory, min_col=2, min_row=1, max_row=len(subcategory_counts) + 1)
create_bar_chart(sheet_subcategory, data_range_subcategory, "Subcategory Distribution", "E2")

sheet_sub_subcategory = wb["Sub-Subcategory Count"]
data_range_sub_subcategory = Reference(sheet_sub_subcategory, min_col=2, min_row=1, max_row=len(sub_subcategory_counts) + 1)
create_bar_chart(sheet_sub_subcategory, data_range_sub_subcategory, "Sub-Subcategory Distribution", "E2")

# Add pie chart for positive and negative feedback
sheet_category = wb["Category Count"]
data_range_category = Reference(sheet_category, min_col=2, min_row=1, max_row=len(category_counts) + 1)
create_pie_chart(sheet_category, data_range_category, "Positive vs Negative Feedback", "E2")

# Save the workbook with charts
wb.save(file_path)

print("Charts successfully added to the Excel file!")
